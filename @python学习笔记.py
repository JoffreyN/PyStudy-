
divmod(a,b) 函数把除数和余数运算结果结合起来，返回一个包含商和余数的元组(a // b, a % b)

itertools迭代{
	from itertools import islice
	for x in islice(iterms, 3, None)	# 迭代，跳过了前面3个元素;如果不要后面的None，则只迭代前面的三个元素
}

tushare金融、股票、电影、票房{
	依赖：pandas、matplotlib、lxml（需下载下来安装）
	tushare.realtime_boxoffice()	#电影票房前十名
	tushare.day_cinema()	#电影院
	tushare.get_realtime_quotes('600355')	#股票当前信息
	tushare.get_hist_data('600355')	#股票历史信息
	tushare.get_latest_news()	#看新闻
	ts.get_stock_basics()	#看股票基本面
	ts.get_sz50s()	获取一些分类信息，如上证50成份股

}

处理pcap{
	from scapy.all import *
	p=rdpcap(path)	#将整个包读取到内存

	from scapy.all import PcapReader
	或
	from scapy.utils import *
	from scapy.route import *
	from scapy.layers.all import *
	packets=PcapReader(path)
	p=packets.read_packet()#一个包一个包的读取，节省内存
}

随机数random{
	random() 返回0<=n<1之间的随机实数n；会生成一个随机的浮点数，范围是在0.0~1.0之间。
	uniform()正好弥补了上面函数的不足，它可以设定浮点数的范围，一个是上限，一个是下限。
	randint()随机生一个整数int类型，可以指定这个整数的范围，同样有上限和下限
	choice(seq) 从序列seq中返回随机的元素；可以从任何序列，比如list列表中，选取一个随机的元素返回，可以用于字符串、列表、元组等
	getrandbits(n) 以长整型形式返回n个随机位；
	shuffle(seq[, random]) 原地指定seq序列；将一个序列中的元素，随机打乱
	sample(seq, n) 从序列seq中选择n个随机且独立的元素；
}


计时{
	import time
	start = time.clock()	//Windows 系统中，建议使用 time.clock()
... do something
	elapsed=time.clock() - start
	
	start = time.time()	# Unix 系统中，建议使用 time.time()
... do something
	elapsed = (time.time() - start)
	
	import timeit
	timeit.timeit("sum(range(100))")
	
	import datetime
	begin = datetime.datetime.now()
... do something
	 elapsed = datetime.datetime.now()-begin
	 
'''
def t():
	start=time.clock()
	exec("from scapy.all import *")
	el=time.clock() - start
	print(el)
'''
}

解码与编码{
	import struct	字节码

	base64{
	import base64
	str(base64.b64encode('abcr34r344r'.encode('utf-8')),'utf-8')	//编码
	str(base64.b64decode('YWJjcjM0cjM0NHI='),'utf-8')	//解码
	}
	
	'你好'.encode('utf-8')	//编码
	b'\xe4\xbd\xa0\xe5\xa5\xbd'.decode('utf-8')	//解码
	b"2018\\u002D02\\u002D07 12:47:41".decode('unicode-escape')	unicode-escape解码

	chardet.detect(b'Hello, world!')	检测编码

	from urllib import parse
	parse.quote(str1)	//url编码
	parse.unquote(str2)	//url解码
}


json{
	json.loads()	//loads是将str转化成dict格式
	json.dumps()		//将dict转化成str格式
	json.dump({'first': 'One', 'second':2}, open('/tmp/result.txt', 'w'))	//dump给的是一个类似于文件指针的东西（并不是真的指针），可以与文件操作结合，而dumps直接给的是str
}

exec语句用来执行储存在字符串或文件中的Python语句。
eval语句用来计算存储在字符串中的有效Python表达式。


获取时间{
	from datetime import datetime
	datetime.now().strftime('%Y-%m-%d %H:%M:%S')	//返回'2017-11-09 09:36:57'

	import time
	time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))	//返回'2017-12-14 11:16:53'
	time.strftime('%H:%M:%S',time.gmtime(120))	秒转化为 时:分:秒

	%y 两位数的年份表示（00-99）
	%Y 四位数的年份表示（000-9999）
	%m 月份（01-12）
	%d 月内中的一天（0-31）
	%H 24小时制小时数（0-23）
	%I 12小时制小时数（01-12） 
	%M 分钟数（00=59）
	%S 秒（00-59）

	%a 简化星期（英语）
	%A 完整星期（英语）
	%b 简化月份（英语）
	%B 完整月份（英语）
	%c 简化星期 简化月 日期 时间 年
	%j 年内的一天（001-366）
	%p 本地A.M.或P.M.的等价符
	%U 一年中的星期数（00-53）星期天为星期的开始
	%W 一年中的星期数（00-53）星期一为星期的开始
	%w 星期（0-6），星期天为星期的开始	
	%x 月/日/两位数的年
	%X 时间
	%Z 当前时区的名称
	%z 时区编号
}

'覆盖之前的打印'{
import sys, time
for progress in range(100):
    time.sleep(0.5)
    sys.stdout.write("Download progress: %d%%   \r" % (progress))
    sys.stdout.flush()

转义符\b是退格键，也就是说把输出的光标往回退格子
转义符\r就可以把光标移动到行首而不换行
转义符\n就把光标移动到行首并且换行
}


openpyxl操作excel{
	写文件
	from openpyxl import Workbook
	wb=Workbook()	
	ws = wb.active
	ws = wb.create_sheet("Mysheet") #插入到最后(default)
	ws = wb.create_sheet(title="Pip", 0) #插入到最开始的位置
	ws.title = "New Title"#修改sheet名字

	excel=Workbook()
	sheet=excel.active
	sheet.title='统计'#修改sheet名字
	sheet2=excel.create_sheet("Epmap统计")#创建第二个sheet
	sheet.append(dicORlist)#按字典写入,可写列表或字典，如果是字典，字典的key必须是1，2，3，4...

	读文件
	from openpyxl import load_workbook
	wb=load_workbook('E:/test.xlsx')
	table = wb.get_sheet_by_name('Sheet1')   #通过表名获取sheet
	  或 table = wb.['Sheet1']	#不知道名字用index
	sheet_names = wb.get_sheet_names()#获取表名
	rows=table.max_row   #获取行数
	cols=table.max_column    #获取列数
	Data=table.cell(row=row,column=col).value  #获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
	table['A1'].value 	#获取表格内容，

	样式
	from openpyxl.styles import *
	c=ws.row_dimensions[1]#获取第一行,填充用此方法获取的单元格对已赋值的单元格无效
	b=ws.column_dimensions['A']#获取第A列,填充用此方法获取的单元格对已赋值的单元格无效
	c=ws['D1']#获取单元格D1
	c=ws[1]#获取第一行中最后一个有数据的单元格之前的所有单元格，
	c.fill=PatternFill(patternType='solid',fgColor='FFFF00')#填充类型为'solid',颜色为黄色
		#先赋值，再填充
	freeze_panes：冻结窗格

	#http://www.mamicode.com/info-detail-2206758.html

}

'xlwt操作excel'{
	excel=xlwt.Workbook()
	sheet1=excel.add_sheet(jsresult['ranks'][0]['billboard_name'])	//创建一个sheet
	sheet=excle.add('sheetname',cell_overwrite_ok=True)	//如果需要对单元格重复操作，需要增加cell_overwrite_ok=True
	sheet1.write(0,1,'Hello Word')	//在第0行第1列写入Hello Word
	excel.save('F:\\QQ部落.xls')	//保存为xls文件
}

'命名变量sheet0、sheet1……'{
names=locals()
for i in range(10):
	names['sheet%s' %i]=i
}

'os库'{
	os.getcwd()	//获取运行的脚本所在目录
	os.path	//获取python所在目录
	os.rename('旧名','新名')	//修改文件名
	os.name字符串指示你正在使用的平台。比如对于Windows，它是'nt'，而对于Linux/Unix用户，它是'posix'。
	os.getcwd()函数得到当前工作目录，即当前Python脚本工作的目录路径。
	os.getenv()和os.putenv()函数分别用来读取和设置环境变量。
	os.listdir()返回指定目录下的所有文件和目录名。
	list(os.walk('E:\\ZP\\Pictures\\tiaoxi\\new'))	 返回[(当前目录，[文件夹]，[文件]),(子目录，[子文件夹]，[子文件]),...]
		遍历目录及其子目录下的所有文件：
		for path,folder,files in os.walk('E:\ZP\Desktop\漏扫相关\openstack_kvm\kvm'):
			for file in files:
				print(os.path.join(path,file))
				
	检验给出的路径是否是一个文件：os.path.isfile()
	检验给出的路径是否是一个目录：os.path.isdir()
	os.remove()函数用来删除一个文件。
	os.system()函数用来运行shell命令。运行成功返回0
	os.linesep字符串给出当前平台使用的行终止符。例如，Windows使用'\r\n'，Linux使用'\n'而Mac使用'\r'。
	os.path.split()函数返回一个路径的目录名和文件名
	os.path.splitext('/path/to/file.txt') 得到文件扩展名('/path/to/file', '.txt')
	os.path.join('/Users/michael', 'testdir')把两个路径合成一个
	os.path.isfile()和os.path.isdir()函数分别检验给出的路径是一个文件还是目录。
	os.path.existe()函数用来检验给出的路径是否真地存在。
	os.path.abspath('.')查看当前目录的绝对路径:
	os.path.dirname(__file__)的使用   http://blog.csdn.net/lxjames833539/article/details/5251608
	os.mkdir('/Users/michael/testdir')创建一个目录
	os.rmdir('/Users/michael/testdir')删掉一个目录
	os.chdir() 方法用于改变当前工作目录到指定的路径。
}

glob.glob(r"E:\Picture\*\*.jpg")  获取指定目录下的所有符合条件的文件

'scrapy框架'{
	scrapy常见命令{
		startproject	//创建一个新工程	scrapy startproject <name> [dir]
		genspider	//创建一个爬虫	scrapy genspider [options] <name><domain>
		settings	//获得爬虫配置信息	scrapy settings [options]
		crawl	//运行一个爬虫	scrapy crawl <spider>
		list	//列出工程中所有爬虫	scrapy list 
		shell	//启动URL调试命令行	scrapy shell [url]
	}
}

'正则表达式'{
	're库主要功能函数'{
		re.search()		//在一个字符串中搜索匹配正则表达式的第一个位置，返回match对象
			{re.search(pattern,string,flags=0)
				pattern:正则表达式的字符串或原生字符串表示
				string:待匹配字符串
				flags:正则表达式使用时的控制标记
					{常用标记:
						re.I	re.IGNORECASE	忽略正则表达式的大小写，
						re.M	re.MULTILINE	正则表达式中的^操作符能将给定字符串的每行当作匹配开始
						re.S	re.DOTALL	正则表达式中的.操作符能匹配所有字符，默认匹配除换行之外的所有字符
					}
			}
		re.match()		//从一个字符串的开始位置其匹配正则表达式，返回match对象
			{'match对象的常见属性'
				.string	//待匹配的文本
				.re	//匹配时使用的pattern对象(正则表达式)
				.pos	//正则表达式搜索文本的开始位置
				.endpos	//正则表达式搜索文本的结束位置
			}
			{'match对象的常见方法'
				.group(0)	//获得匹配后的字符串
				.start()	//匹配字符串在原始字符串的开始位置
				.end()	//匹配字符串在在原始字符串的结束位置
				.span()	//返回(.start(),.end())
			}	
				
		re.findall()		//搜索字符串，以列表类型返回全部能匹配的子串
		re.split()		//将一个字符串按照正则表达式匹配结果去掉，把剩余部分进行分割，返回列表类型
			re.split(pattern,string,maxsplit=0,flags=0)
				maxsplit:最大分割数
		re.finditer()		//搜索字符串，返回一个匹配结果的迭代类型，每个迭代元素是match对象
		re.sub()		//在一个字符串中替换所有匹配正则表达式的子串，返回替换后的字符串
			{re.sub(pattern,repl,string,count=0,flags=0)
				repl:替换匹配字符串的字符串
				count:匹配的最大替换次数}
	}
	
	're库的两种等价用法'{
		函数式用法：>>>rst=re.search(r'[1-9]\d{5}','BIT 100081')
		面向对象用法：
			>>>pat=re.compiler(r'[1-9]\d{5}')
			>>>rst=pat.search('BIT 100081')
	}
	
	'常用操作符'{
		.		表示任何单个字符
		[]	字符集，对单个字符给出取值范围	[abc]表示a、b、c，[a-z]表示a到z单个字符
		[^ ]	非字符集，对单个字符给出排除范围	[^abc]表示非a或b或c的单个字符
		*	前一个字符0次或无限次扩展	abc*表示ab,abc,abcc,abccc等	*?,最小匹配
		+	前一个字符1次或无限次扩展	abc+表示abc,abcc,abccc等	+?,最小匹配
		?	前一个字符0次或1次扩展	abc?表示ab,abc	??,最小匹配
		|	左右表达式任意一个	abc|def表示abc,def
		{m}		扩展前一个字符m次	ab{2}c表示abbc
		{m,n}	扩展前一个字符m至n次	ab{1,2}c表示abc、abbc	{m,n}?,最小匹配
		^	匹配字符串开头	^abc表示abc且在一个字符串的开头
		$	匹配字符串结尾	abc$表示abc且在一个字符串的结尾
		()	分组标记，内部只能使用|操作符		(abc)表示abc,(abc|def)表示abc、def
		^(\d{3})-(\d{3,8})$分别定义了两个组，可以直接从匹配的字符串中提取出区号和本地号码
		\d	数字，等价于[0-9]
		\w	单词字符，等价于[A-Za-z0-9]
		\s可以匹配一个空格（也包括Tab等空白符）
		? 非贪婪匹配
	}
	
	实例{
		^[A-Za-z]+$		由26个字母组成的 字符串
		^[A-Za-z0-9]+$	有26个字母和数字组成的字符窜
		^-?\d+$		整数形式的字符串
		^[0-9]*[1-9][0-9]*$		正整数形式的字符串
		[1-9]\d{5}	中国境内邮政编码，6位
		[\u4e00-\u9fa5]		匹配中文字符
		\d{3}-\d{8}|\d{4}-\d{7}	国内电话号码，010-68913536
		[1-9]?\d	0-99
		1\d{2}	100-199
		2[0-4]\d	200-249
		25[0-5]		250-255
		2[0-5][0-5]
		re.match(r'^(\d+?)(0*)$', '102300').groups()	非贪婪匹配返回('1023', '00')
	}
}

'BeautifulSoup库'{
	<>.find_all(name,attrs,recursive,string,**kwargs)	//返回一个列表类型，存储查找结果
		name:对标签名称的检索字符串
		attrs:对标签属性值的检索字符串，可标注属性检索
		recursive:布尔型，是否对子孙全部检索，默认True
		string:检索标签内容
	data=soup.find_all('span',{'class':'c-gap-right'})	//查找class=c-gap-right的所有span标签
	
	from bs4 import BeautifulSoup	//导入
	demo=requests.get(url).text
	soup=BeautifulSoup(demo,"html.parser")	//解析一个网页，demo为通过requests.get获得的网页，"html.parser"为html解析器，
	soup=BeautifulSoup(open("D://demo.html"),"html.parser")	//打开本地html
	soup.prettify()
	
	BeautifulSoup类的基本元素{
		Tag		//标签，最基本的信息组织单元，分别用<></>标明开头和结尾
		Name		//标签的名字，格式：T.name
		Attributes		//标签的属性，字典形式组织，格式：T.attrs
		NavigableString		//标签内容，格式：T.string
		Comment		//标签内字符串的注释部分，一种特殊的Comment类型
		parent	//获取该元素的父元素，T.a.parent.name
	}
	
	标签树的下行遍历{
		.contents	//子节点的列表，将<tag>所有子节点存入列表，包含字符串
		.children	//子节点的迭代类型，与.contents类似，用于循环遍历子节点
		.descendants	//子孙节点的迭代类型，包含所以子孙节点，用于循环遍历
	}
	
	标签树的上行遍历{
		.parent		//节点的父节点
		.patents	//节点的父爷节点的迭代类型，用于循环遍历先辈节点
	}
	
	标签的平行遍历{
		.next_sibling		//返回按照HTML文本顺序的下一个平行节点标签或标签内容
		.previous_sibling		//返回按照HTML文本顺序的上一个平行节点标签或标签内容
		.next_siblings		//迭代类型，返回按照HTML文本顺序的后续所有平行节点标签或标签内容
		.previous_siblings		//迭代类型，返回按照HTML文本顺序的前续所有平行节点标签或标签内容
	}
}

'requests库'{
http://docs.python-requests.org/zh_CN/latest/user/quickstart.html
	r = requests.get(url)
	r.headers	获取响应的headers
	r.json 	对于特定类型的响应，例如JSON，可以直接获取
	r.raw	原始响应内容（r.raw.read(10)）

	requests.request()		//构造一个请求，支撑以下各方法的基础方法
	requests.get()		//获取HTML网页的主要方法，对应于HTTP的GET
	requests.head()		//获取HTML网页头信息的方法，对应于HTTP的head
	requests.post()		//向HTML页面提交POST请求的方法，对应于HTTP中的POST
		requests.post(url,headers=header,data=key)	key为要post的数据，格式可以为字典，也可以是key=json.dumps(key)
		#https://blog.csdn.net/junli_chen/article/details/53670887	Python 使用requests发送POST请求
	requests.put()		//向HTML页面提交PUT请求的方法，对应于HTTP中的PUT
	requests.patch()		//向HTML页面提交局部修改请求，对应于HTTP中的PATCH
	requests.delete()		//向HTML页面提交删除请求，对应于HTTP中的DELETE
	
	requests.ConnectionError		//网络连接错误异常。如DNS查询失败、拒绝连接等
	requests.HTTPError		//HTTP错误异常
	requests.URLRequired		//URL缺失异常
	requests.TooManyRedireets		//超过最大重定向次数，产生重定向异常
	requests.ConnectTimeout		//连接远程服务器超时异常
	requests.Timeout		//请求URL超时，
	
	requests.request(method,url,**kwargs)	//e.g:requests.request('GET','http://example.com',data=key1)
		**kwargs:控制访问的参数，可选，有很多个{
			params:字典或字节序列，作为参数增加到URL中
			data:字典、字节序列或文件对象，作为Request的内容
			json:JSON格式的数据，作为request的内容
			heads:字典，http定制头
			cookies:字典或CookieJar，Request中的cookie
			auth:元组，支持HTTP认证功能
			files:字典类型，传输文件
			timeout:设定超时时间，单位秒
			proxies:字典类型，设定访问代理服务器，可增加登录认证。可防止对爬虫的逆追踪
			allow_redirects:True/False,默认为True，重定向开关
			stream:True/False,默认为True，获取内容立即下载开关
			verify:True/False,默认为True，认证SSL证书开关
			cert:本地SSL证书路径
		}

}

图像处理{
http://www.cnblogs.com/apexchu/p/4231041.html
	from PIL import Image
	Image.open(path)	打开图片
}

文件处理{
	for line in open('myfile.txt','r'):
		print(line, end='')

	shutil.copyfile('a.py', 'copy_a.py')	复制文件

	import fileinput
	for i in fileinput.input('E:\\ZP\\Desktop\\新建文本文档.txt'):
		print i
	
	http://blog.csdn.net/scelong/article/details/6971917
	<variable>=open(<name>,<mode>)	//<variable>变量名；<name>文件名；<mode>打开模式
	infile=open("123.txt","r")
	f = open('test.txt', 'r', encoding='gbk', errors='ignore')	//指定编码，并忽略错误
	infile.readlines()
	read()方法可以一次读取文件的全部内容
	read(size)方法，每次最多读取size个字节的内容。
	readline()可以每次读取一行内容
	readlines()一次读取所有内容并按行返回list
	f.write(str)		//写入字符串
	f.writelines(list)		//写入列表
	<mode>{
		r 	//只读，如果文件不存在则报错
		w 	//只写，如果文件不存在则创建文件,如果有，那么就会先把原文件的内容清空再写入新的东西。
		a 	//附加到文件末尾,如果文件不存在则创建文件
		rb	//只读二进制文件，如果文件不存在则报错
		wb	//只写二进制文件，如果文件不存在则创建文件
		ab	//附加到二进制文件末尾
		r+	//r+w（可读可写，文件若不存在就报错(IOError)）
		w+ // w+r（可读可写，文件若不存在就创建）
		a+ //a+r（可追加可写，文件若不存在就创建）
	}
}

列表的操作{
	re.findall(r'\[(.+?)\]',text)	提取字符串中的列表
	[对(x)的操作 for x in 集合 if 条件]
	[满足条件后对i操作 if 条件 else 不满足条件后对i操作 for i in 集合]
	[对(x,y)的操作 for x in 集合1 for y in 集合2 if 条件]
		例：a,b,c=[1,2,3],[4,5,6],[7,8,9]
			[[x,y,z] for x in a for y in b for z in c if a.index(x)==b.index(y)==c.index(z)]	返回[[1, 4, 7], [2, 5, 8], [3, 6, 9]]
			['4' if x == '1' else x for x in a]	将a中的1替换为4
			[dic[x] if x in dic else x for x in list]
	<list>.append(x)	//将元素x增加到列表最后
	<list>.sort()	//将列表元素排序
	items.sort(key=lambda x:x[1],reverse=True)	//参数key表示排序的项，key=lambda x:x[1]表示按list中索引为1的元素排序；rever=True表示降序，反之则升序
	key=lambda x:x+1	//快速定义函数，冒号前为行参，后为函数表达式；输入key(2)则返回3
	<list>.reverse()	//将列表元素反向排序，即最后一个排到第一个
	<list>.index(x)	//返回第一次出现元素x的索引值
	<list>.insert(i,x)	//在位置i出插入新元素x
	<list>.count(x)	//返回元素x在列表中的数量
	set(<list>)	 使用set函数可删除列表中的重复项，去重
		set() 函数创建一个无序不重复元素集，可进行关系测试，删除重复数据，还可以计算交集、差集、并集等。
	
	<list>.remove(x)	//删除列表中第一次出现的元素x
	'-'.join(<list>)	//将字符串、元组、列表中的元素以字符'-'连接生成一个新的字符串
	<list>.pop(i)	//取出列表中位置为i的元素，并删除
	list1 = [2, 3, 4]
	list2 = [2*i for i in list1 if i > 2]	//得到一个新列表，使list1中所有大于2的数都是原来的2倍。
	deque除了实现list的append()和pop()外，还支持appendleft()和popleft()，这样就可以非常高效地往头部添加或删除元素。
	sum([[1, 10], [2, 9], [3, 8], [4, 7], [5, 6]],[])	合并为大列表[1, 10, 2, 9, 3, 8, 4, 7, 5, 6]
	sum([['asdfsa','asf','asf'],['asfda','fsdf','fsdf']],[])	合并为大列表['asdfsa', 'asf', 'asf', 'asfda', 'fsdf', 'fsdf']
	
	列表生成{
	>>> [x * x for x in range(1, 11)]	//生成[1, 4, 9, 16, 25, 36, 49, 64, 81, 100]；(1,11)表示从1开始到11结束，包含1不包含11
	>>> [x * x for x in range(1, 11) if x % 2 == 0]	//筛选出仅偶数的平方：[4, 16, 36, 64, 100]
	>>> [m + n for m in 'ABC' for n in 'XYZ']	//使用两层循环，可以生成全排列：['AX', 'AY', 'AZ', 'BX', 'BY', 'BZ', 'CX', 'CY', 'CZ']
	[x for x in range(1,254) if x!=170 ]
	>>> import os	//导入os模块
	>>> [d for d in os.listdir('.')]	//列出当前目录文件夹内容

	>>> d = {'x': 'A', 'y': 'B', 'z': 'C' }	//定义字典d
	>>> [k + '=' + v for k, v in d.items()]	//列表生成式，使用两个变量来生成list：['y=B', 'x=A', 'z=C']

	>>> L = ['Hello', 'World', 'IBM', 'Apple']
	>>> [s.lower() for s in L]	//把一个list中所有的字符串变成小写;upper()则相反
	choice(L)	//在列表中随机选择一个数

}
}

集合set{与列表和元组不同，集合是无序的，也无法通过数字进行索引。此外，集合中的元素不能重复。
https://www.cnblogs.com/9527chu/p/5520752.html
	set() 函数创建一个无序不重复元素集，可进行关系测试，删除重复数据，还可以计算交集、差集、并集等。
	set(<list>)	 使用set函数可删除列表中的重复项，
	a = set('boy')	定义一个集合
	b=set('python')
	a.add('python') 把要传入的元素做为一个整个添加到集合中
	a.update('python')	把要传入的元素拆分，做为个体传入到集合中
	a.remove('python')	移除集合中的某个元素，如果该元素不再集合中，则报错
	s1.discard(22)	移除集合中的某个元素，如果该元素不在集合中，则什么也不做。
	a&b 	交集
	s.intersection(s1)	返回两个集合的交集
	a|b 	并集
	a-b 	差集，在集合a中减去b中有的
	s.clear()	清空集合中的所有元素
	s1=s.copy()	获得一个集合的副本，并且属于浅拷贝。
	A.difference(B)	在集合A中但不在集合B中
	A.difference_update(B) 	类似difference，区别是将结果复制给A.
	
}

tuple元组{

	namedtuple是一个函数，它用来创建一个自定义的tuple对象，并且规定了tuple元素的个数，并可以用属性而不是索引来引用tuple的某个元素。
}

排列组合{
	import itertools	http://python.jobbole.com/87455/
	list(itertools.permutations([1,2,3,4],2))	排列
	list(itertools.combinations([1,2,2,4],2))	组合
	itertools.count(1,2) 	创建一个无限迭代器，从1开始，步进2
}

字典{
dict(a="1", b="2")	//返回一个字典{'a': '1', 'b': '2'}
d={'Michael':98,'Bob':83,'Tracy':89}	//定义一个dict
d['Adam']=67	//在名称为d的dict中插入数据
del d['Adma']	//删除key为Adma的数据
'thomas' in d	//检查d中是否存在Thomas，返回false则不存在
d.get('Thomas')	//检查d中是否存在Thomas，返回none则不存在;注意：交互式命令行不显示结果。
d.keys()	//返回字典中所有的key
d.values()	//返回所有value
d.items()	//返回所有键值的列表
d.pop('Adma')	//删除对应的值
d.clear()	//删除所有
d.update(t)	//将字典t中的键值添加到字典d中
dict(zip(['one', 'two', 'three'], [1, 2, 3]))  生成字典

和list比较，dict有以下几个特点：
    查找和插入的速度极快，不会随着key的增加而变慢；
    需要占用大量的内存，内存浪费多。

而list相反：
    查找和插入的时间随着元素的增加而增加；
    占用空间小，浪费内存很少。

    dd = defaultdict(lambda: 'N/A')如果希望key不存在时，返回一个默认值，就可以用defaultdict
    od = OrderedDict([('a', 1), ('b', 2), ('c', 3)])如果要保持Key的顺序，可以用OrderedDict，按照插入的顺序排列

}   

常见字符串处理函数{
	str.rindex(substr,start,end)		//rindex()函数类似于rfind()函数，在Python中也是在字符串中倒序查找子串最后一次出现的位置，跟rfind()不同的是，未找到则抛出异常。
	str.count(substr,start,end)	//在字符串str中统计子串substr出现的次数，如果不指定开始位置start和结束位置end，表示从头统计到尾。
	str.expandtabls(tabsize)	//将str字符串中的tab字符替换成tabsize个空格，默认是8个空格。
	<string>.upper()	//将字符串中字母转换为大写
	<string>.lower()	//将字符串中字母转换为小写
	<string>.capitalize()	//将字符串首字母大写，其余小写
	<string>.strip()	//去掉两边的指定字符
	<string>.rstrip('\n')	//去掉尾部的'\n'
	<string>.split()	//按指定字符分割字符串为列表，指定的字符不会显示
	<string>.replace("\n", "")	//替换字符串中的所有'\n'
	jieba.lcut("中国是个伟大的国家")	//中文分词函数
	<string>.isdigit()	//判断是否为数字类型，是则返回True
	<string>.find("xx")	//查找字符串中出现xx的位置
	<string>.count("xx")	//查找字符串中包含xx的字符串个数
	<string>.replace()	//字符串替换
	<string>.replace(' ',':',1)	//将字符串中的空格替换为':'，只替换1次
	'--help'.startswith('--')		//判断字符串是否以'--'开始
	'--help'.endswith('--')		//判断字符串是否以'--'结束
	'123456'[::-1]	//字符串逆序输出，反向，翻转。反转
	<string>.isdigit()	//判断字符串是否为纯数字
	<string>.isalpha()	//判断字符串是否为纯字母
	<string>.isalnum()	//判断字符串是否为数字字母组合
	<string>.isspace()	//判断字符串是否为空格

	python:字符串转换成字节的三种方式{
		b'zifuchuang'
		bytes('zifuchuang',encoding='utf-8')
		('zifuchuang').encode('utf-8')
		}
}	

divmod(x,y)	//等于(x//y,x%y)

type(x)	//返回x数据类型

import 库名	//引用库，调用函数时需要使用库名，如turtle.fd(10)
from 库名 import 函数名	//引用库，调用函数时则不需使用库名，如fd(10)
from 库名 import *

print(4>>1)  4除以2
print(4<<1)  4乘以2



map(f,[1,2,3])将传入的函数依次作用到序列的每个元素，并把结果作为新的Iterator返回。
reduce(f,[1,2,3])把结果继续和序列的下一个元素做累积计算
filter(f, [1,2,3])把传入的函数依次作用于每个元素，然后根据返回值是True还是False决定保留还是丢弃该元素。
	把一个序列中的空字符串删掉，可以这么写：
	def not_empty(s):
    return s and s.strip()
  list(filter(not_empty, ['A', '', 'B', None, 'C', '  ']))
sorted()函数也是一个高阶函数，它还可以接收一个key函数来实现自定义的排序，例如按绝对值大小排序  
	sorted([36, 5, -12, 9, -21], key=abs)
	第三个参数reverse=True表示反向排序
nonlocal	//在 Python 中，内层函数对外层作用域中的变量仅有只读访问权限！而 nonlocal 可以使我们自由地操作外层作用域中的变量！
functools.partial	的作用就是，把一个函数的某些参数给固定住（也就是设置默认值），返回一个新的函数，调用这个新函数会更简单。
isinstance([1, 2, 3], (list, tuple))		//判断一个变量是否是某些类型中的一种
hasattr(obj, 'power') # 有属性'power'吗？
getattr(obj, 'power') # 获取属性'power'
__slots__		//在定义class的时候，定义一个特殊的__slots__变量，来限制该class实例能添加的属性
 
 
  
赋值语句1：
a, b = b, a + b

相当于：
t = (b, a + b) # t是一个tuple
a = t[0]
b = t[1]	//但不必显式写出临时变量t就可以赋值。

赋值语句2：
x,y=y,x		//交换x,y的值

迭代{	//for循环来遍历这个list或tuple，这种遍历我们称为迭代（Iteration）。
d={'a':1,'b':2,'c':3,'d':4}	//字典d
for key in d:print(key)	//迭代key
for value in d.values()	//迭代values
for k, v in d.items()	//同时迭代key和value
for ch in 'ABC':print(ch)	//迭代字符串

from collections import Iterable	//使用collections模块的Iterable类型
isinstance('abc',Iterable)	//判断‘abc’是否可迭代，返回true则可以

for i,value in enumerate(['A','B','C','D']):print(i,value)	//enumerate函数可以把一个list变成索引-元素对，这样就可以在for循环中同时迭代索引和元素本身

import itertools
count()会创建一个无限的迭代器，步进2:  for n in itertools.count(0,2)
cycle()会把传入的一个序列无限重复下去:  for c in itertools.cycle('ABC')
repeat()负责把一个元素无限重复下去，不过如果提供第二个参数就可以限定重复次数:  for n in itertools.repeat('A', 3)
takewhile()函数根据条件判断来截取出一个有限的序列:  ns=itertools.takewhile(lambda x: x <= 10, itertools.count(1))	 
chain()可以把一组迭代对象串联起来，形成一个更大的迭代器:  for c in itertools.chain('ABC', 'XYZ')
groupby()把迭代器中相邻的重复元素挑出来放在一起:  for key, group in itertools.groupby('AAABBBCCAAA'):print(key, list(group))
	for key, group in itertools.groupby('AAABBBCCAAAaa',lambda c: c.upper()):print(key, list(group))	不区分大小写

}



生成器generator{
创建generator方法1：
g = (x * x for x in range(10))	//创建List和generator的区别仅在于最外层的[]和().
>>> next(g)	//可以通过next()函数获得generator的下一个返回值

创建generator方法2：


}

在 面向过程 的语言中，程序是由过程或仅仅是可重用代码的函数构建起来的。
在 面向对象 的语言中，程序是由数据和功能组合而成的对象构建起来的。

一定要用自然字符串处理正则表达式。

如赋值运算符那样的运算是由右向左运算，如a=b=c等同于a=(b=c)


len('中文'.encode('utf-8'))	//计算字节数；结果为6

#!/usr/bin/env python3	//告诉Linux/OS X系统，这是一个Python可执行程序，Windows系统会忽略这个注释；
# -*- coding: utf-8 -*-	//告诉Python解释器，按照UTF-8编码读取源代码，否则在源代码中写的中文输出可能会有乱码





classmate=('Michael', 'Bob', 'Tracy')	//定义tuple列表；无法修改。
t = (1,)	//定义只有1个元素的tuple列表t

range(101)	//生成0~100的整数数列

 
s = set([1, 2, 3])	//创建一个set，set和dict类似，也是一组key的集合，但不存储value，没有重复的key。
s.add(4)	//添加元素到set中
s.remove(4)	//删除元素

isinstance(x, (int, float))	//判断x的值是否为整数或浮点数

import math	//表示导入math包，并允许后续代码引用math包里的sin、cos等函数。

def calc(*numbers)	//定义一个函数calc，其中参数数量可变
calc(*nums)	//参数为一个list或tuple

def person(name, age, **kw)	//定义一个函数person，其中参数kw为关键字参数；关键字参数在函数内部自动组装为一个dict。
person('Jack', 24, **extra)	//**extra表示把extra这个dict的所有key-value用关键字参数传入到函数的**kw参数，kw将获得一个dict，注意kw获得的dict是extra的一份拷贝，对kw的改动不会影响到函数外的extra。
def person(name, age, *, city, job)	//只接收city和job作为关键字参数。传入参数时必须加关键字参数名
def person(name, age, *args, city, job)	//其中args为可变参数，且只接收city和job作为关键字参数。
参数定义的顺序必须是：必选参数、默认参数、可变参数、命名关键字参数和关键字参数。

L[:10]		//取出前10个数
L[-10:]		//取出后10个数
L[10:20]		//取出前11-20个数
L[:10:2]		//取出前10个数，每两个取一个
L[::5]		//取出所有数，每5个取一个
对tuple、字符串同样有效

*args，**kwargs 这两个参数	http://blog.csdn.net/callinglove/article/details/45483097